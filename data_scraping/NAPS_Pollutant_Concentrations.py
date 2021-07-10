from Postgres import Postgres
from Tables import Tables
from urllib.request import Request, urlopen
import requests
import zipfile
import shutil
from bs4 import BeautifulSoup
import os
from datetime import date, datetime
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from typing import Union, List, Dict, Tuple

# TODO: try sqlalchemy

class NAPS_Pollutant_Concentrations():

    FIRST_YEAR = 1990
    LAST_YEAR = datetime.now().year - 1 # TODO: check this range

    INTEGRATED_POLLUTANTS = ["CARBONYLS", "VOC", "PAH", "PM2.5", "PM2.5-10"]
    

    def __init__(self, hostname, database, user, password) -> None:
        
        self._psql = Postgres(hostname, database, user, password)

        self._create_tables()


    def update_data(self, year: int=None) -> None:

        self._update_continuous_data(year=year)
        self._update_integrated_data(year=year)


    def _update_continuous_data(self, year: int=None) -> None:

        if year:
            self._get_data(year, 0)
        else:
            for year in self._get_valid_year_range(): # starts from first year
                self._get_data(year, 0)

    def _update_integrated_data(self, year: int=None) -> None:

        if year:
            self._get_data(year, 1)
        else:
            for year in self._get_valid_year_range()[::-1]: # starts last year and goes backwards
                self._get_data(year, 1)

    def _create_tables(self) -> None:

        Tables.connect(self._psql)
        Tables.create_naps_continuous()
        
        for pollutant in NAPS_Pollutant_Concentrations.INTEGRATED_POLLUTANTS:
            Tables.create_naps_integrated_pollutant(pollutant)

    def _get_data(self, year: int, dataTypeVal: int) -> None:

        # dataTypeVal = 1 -> IntegratedData
        # dataTypeVal = 0 -> continuousData

        column_order = {} # {column: index}

        # connect to webpage
        dataType = "IntegratedData-DonneesPonctuelles" if dataTypeVal else "ContinuousData-DonneesContinu/HourlyData-DonneesHoraires"
        url = f"https://data-donnees.ec.gc.ca/data/air/monitor/national-air-pollution-surveillance-naps-program/Data-Donnees/{year}/{dataType}/?lang=en"
        request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        page = urlopen(request)
        html = page.read().decode("utf-8")
        soup = BeautifulSoup(html, "html.parser")

        # confirm relevant column indices
        th = soup.find_all("th")
        for i in range(len(th)):
            if th[i].text in ("Name", "Description"):
                column_order[th[i].text] = i

        # get relevant rows of data
        tr = soup.find_all("tr")
        for i in range(len(tr)):
            if tr[i]["class"][0] != "indexhead": # ignore row from header
                td = tr[i].find_all("td")
                description = "Data archive" if dataTypeVal else "Comma separated"

                if td[column_order["Description"]].text == description:
                    filename = td[column_order["Name"]].text
                    if "ReferenceMethod" not in filename:
                        filepath = f"{os.path.dirname(os.path.realpath(__file__))}\\{filename}"
                        file_url = url.split('?')[0] + filename
                        zip_dir = self.__download_file(file_url, filepath)

                        if dataTypeVal:
                            data = self._get_integrated_data(zip_dir)
                            self._insert_integrated_data(data)
                        else:
                            data = self._get_continuous_data(filepath) 
                            self._insert_continuous_data(data)
                        self.__delete_files(filepath, zip_dir)

    def _insert_continuous_data(self, data: List[Dict[str, str]]) -> None:

            pollutant = data[0]["Pollutant"] # should only have one pollutant per csvfile
            pollutant_id = Tables.get_naps_pollutants(pollutant)
            
            with tqdm(data) as tlines:
                for line in tlines:
                    naps_id = line["NAPSID"]
                    naps_station_id = Tables.get_naps_station(naps_id)
                    date = line["Date"]
                    year, month, day = int(date[:4]), int(date[4:6]), int(date[6:])
                    tlines.set_description(f"Continuous-{pollutant}-{year} Uploading")
                    
                    for hour in range(24):
                        timestamp = str(datetime(year, month, day, hour))
                        density = "NULL" if line[hour] is None else float(line[hour])
                        command = f"""
                            INSERT INTO {Tables.NAPS_CONTINUOUS} (timestamp, naps_station, pollutant, density)
                            VALUES (%(timestamp)s, {naps_station_id}, {pollutant_id}, {density})
                        """
                        str_params = {"timestamp": timestamp}
                        self._psql.command(command, 'w', str_params=str_params)
    
    def _get_continuous_data(self, filepath: str) -> List[Dict[str, str]]:

        pollutant = filepath.split('\\')[-1].split('_')[0]
        year = filepath.split('\\')[-1].split('_')[-1].split('.')[0]
        
        column_order = {}
        total_data = []
        data_len = 31

        df = pd.read_csv(filepath, sep='\n', error_bad_lines=False)
        for row in tqdm(range(len(df)), desc=f"Continuous-{pollutant}-{year} Formatting"):
            line = df.loc[row][0].split(',')
            if column_order or (len(line) == data_len and line.count('') == 0): # first row is column headers
                if not column_order:
                    for i in range(data_len):
                        if len(line[i].split("//")[0]) == 3 and line[i].split("//")[0][0] == 'H': # remove "H0" from hour
                            column_order[i] = int(line[i].split("//")[0].replace('H', '').replace('24', '0')) # TODO: is H24 midnight?
                        else:
                            column_order[i] = line[i].split("//")[0]
                
                else:
                    data = {}
                    for i in range(data_len):
                        data[column_order[i]] = line[i]
                    total_data.append(data)

        return total_data

    def _insert_integrated_data(self, data: Tuple[str, List[Dict[str, str]]]) -> None:

        pollutant, workbooks = data
        table = Tables.integrated_pollutants[pollutant.upper()]["table"]
        compounds = Tables.integrated_pollutants[pollutant.upper()]["compounds_table"]
        total_lines = sum([sum([len(wb[sheet]) for sheet in wb.keys()]) for wb in workbooks.values()])

        with tqdm(total=total_lines) as pbar:
            for naps_id in workbooks:
                naps_station_id = Tables.get_naps_station(int(naps_id))
                for sheet in workbooks[naps_id]:
                    for line in workbooks[naps_id][sheet]:
                        sample_type_id = Tables.get_naps_sample_type(line.pop("Sample Type"))
                        year, month, day = [int(d) for d in line.pop("Sampling Date").split('-')] # expecting 'year-month-day'
                        sampling_date = str(date(year, month, day))
                        pbar.set_description(f"Integrated-{pollutant}-{year} Uploading")

                        for compound in line.keys():
                            if "MDL" not in compound and "VFlag" not in compound:
                                compound_name, metadata = compound.split("-metadata:")
                                compound_id = Tables.get_naps_integrated_pollutant_compound(pollutant.upper(), compound)
                                density = "NULL" if line[compound] is None else float(line[compound])

                                if "Recovery" in compound_name:
                                    has_mdl, has_vflag = False, False
                                elif pollutant == "PM2.5" and sheet == "OCEC & OC Artifacts" and ("POC" in compound_name or [c for c in compound_name if c.isdigit()]):
                                    has_mdl = False
                                else:
                                    has_mdl, has_vflag = True, True

                                # find short name if available
                                # example: 'Acenaphthylene (AL)' => 'AL'
                                # does not apply in this specific sheet: 'OCEC & OC Artifacts' for PM2.5
                                if compound_name[-1] == ')' and not (pollutant == "PM2.5" and sheet == "OCEC & OC Artifacts"):
                                    compound_name = compound_name.split()[-1][1:-1]
                                
                                if has_mdl:
                                    density_mdl_name = f"{compound_name}-MDL-metadata:{metadata}"
                                    density_mdl = "NULL" if line[density_mdl_name] is None else float(line[density_mdl_name])
                                else:
                                    density_mdl = "NULL"

                                if has_vflag:
                                    vflag_name = f"{compound_name}-VFlag-metadata:{metadata}"
                                    vflag = "NULL" if (not line[vflag_name] or not line[vflag_name].strip()) else line[vflag_name]
                                    vflag_id = Tables.get_naps_validation_code(vflag)
                                else:
                                    vflag_id = "NULL"

                                command = f"""
                                    INSERT INTO {table} (sampling_date, naps_station, sample_type, compound, density, density_mdl, vflag)
                                    VALUES (%(sampling_date)s, {naps_station_id}, {sample_type_id}, {compound_id}, {density}, {density_mdl}, {vflag_id})
                                """
                                str_params = {"sampling_date": sampling_date}
                                self._psql.command(command, 'w', str_params=str_params)
                        
                        pbar.update(1)

    def _get_integrated_data(self, zip_dir: str) -> Tuple[str, List[Dict[str, str]]]:

        pollutant = "PM2.5-10" if zip_dir.split('\\')[-1] == "PM2.5-10" else zip_dir.split('\\')[-1].split('-')[0]

        unwanted_sheets = ["Station Info", "Metadata Information"]
        # can get naps id from file name
        unwanted_columns = ["NAPS Site ID", "NAPS ID", "Start Time", "End Time", "Actual Volume", "Pres.", "Temp."]

        total_data = {}
        with tqdm(os.listdir(zip_dir)) as tfiles:
            for f in tfiles:
                if f.split('.')[-1] == "xlsx" and f.split('.')[0].split('_')[-1] == "EN":
                    year = f.split('_')[2]
                    naps_id = f.split('_')[0].replace('S', '')
                    xlsx_filepath = f"{zip_dir}\\{f}"
                    wb = load_workbook(xlsx_filepath)
                    tfiles.set_description(f"Integrated-{pollutant}-{year} Formatting")
                    tfiles.set_postfix(naps_id=naps_id)
                    wb_data = {}

                    for sheet_name in [name for name in wb.sheetnames if name not in unwanted_sheets]:
                        sheet = wb[sheet_name]
                        column_order = {} # col: name
                        meta_data_headers = ["Medium", "Observation Type", "Analytical Instrument", "Analytical Method", "Cartridge", "Units"]
                        meta_data_rows = {} # name: row
                        data_len = None
                        sheet_data = []

                        for row in range(1, sheet.max_row+1):
                            # find column headers
                            if not column_order:
                                # find metadata headers
                                if sheet[row][0].value in meta_data_headers:
                                    meta_data_name = sheet[row][0].value
                                    meta_data_name = meta_data_name.replace("Analytical Method", "Analytical Instrument")
                                    meta_data_rows[meta_data_name] = row
                                    
                                line = ["None" if not cell.value else str(cell.value) for cell in sheet[row]]
                                if line[0] in ["NAPS Site ID", "NAPS ID"]: # TODO: find better way than hardcoding column header start
                                    line = list(filter(lambda val: val != "None", line)) # strip None values from end of row
                                    data_len = len(line)

                                    i = 0
                                    while i < data_len:
                                        meta_data = {}
                                        for meta_data_name in meta_data_rows.keys(): # only use meta_data that was found
                                                meta_data_val = sheet[meta_data_rows[meta_data_name]][i].value
                                                if meta_data_val:
                                                    meta_data_val = meta_data_val.replace("ICPMS", "ICP/MS")
                                                    meta_data_val = meta_data_val.replace('GC-MS', 'GC/MS')
                                                    meta_data_val = meta_data_val.replace('_', '/')
                                                    meta_data_val = meta_data_val.replace("N/A", "NULL")
                                                    meta_data_val = meta_data_val.replace('µ', 'u')
                                                    meta_data_val = meta_data_val.replace(',', '+')
                                                    meta_data_val = ''.join(meta_data_val.split()) # remove whitespace
                                                    meta_data[meta_data_name] = meta_data_val
                                                else:
                                                    meta_data[meta_data_name] = "NULL"

                                        value = sheet[row][i].value
                                        if value in unwanted_columns:
                                            i += 1
                                        elif value in ["Sampling Type", "Sample Type", "Sampling Date", "Sample Date"]:
                                            value = value.replace("Sampling Type", "Sample Type")
                                            value = value.replace("Sample Date", "Sampling Date")
                                            column_order[i] = value
                                            i += 1
                                        elif "Recovery" in value:
                                            column_order[i] = f"{value}-metadata:{meta_data}"
                                            i += 1
                                        else:
                                            # this specific sheet 'OCEC & OC Artifacts' in PM2.5 doesn't have MDL for some columns
                                            num_cols = 2 if (pollutant == "PM2.5" and sheet_name == "OCEC & OC Artifacts" and ("POC" in value or [c for c in value if c.isdigit()])) else 3
                                            for j in range(num_cols):
                                                value = sheet[row][i+j].value
                                                value = value.replace("Vflag", "VFlag")
                                                column_order[i+j] = f"{value}-metadata:{meta_data}"
                                            i += num_cols
                            
                            else:
                                line_data = {}
                                for i in column_order:
                                    # convert datetime object to string 'year-month-day' for consistency
                                    if type(sheet[row][i].value) == datetime:
                                        line_data[column_order[i]] = str(sheet[row][i].value).split()[0]
                                    else:
                                        line_data[column_order[i]] = sheet[row][i].value
                                sheet_data.append(line_data)
                        wb_data[sheet_name] = sheet_data     
                    total_data[naps_id] = wb_data
    
        return pollutant, total_data

    def _get_valid_year_range(self) -> List[int]:

        # changes const value instead of function output in case it returns None
        start_year = self._get_most_recent_year() or NAPS_Pollutant_Concentrations.FIRST_YEAR-1 
        year_range = [year for year in range(start_year+1, NAPS_Pollutant_Concentrations.LAST_YEAR)]

        return year_range
        
    def _get_most_recent_year(self) -> Union[int, None]:

        cmd = f"""
            SELECT timestamp FROM {Tables.NAPS_CONTINUOUS} ORDER BY timestamp DESC LIMIT 1 
        """
        timestamp = self._psql.command(cmd, 'r')

        return None if timestamp == [] else timestamp[0][0].year

    def __download_file(self, url: str, filepath: str) -> Union[str, None]:

        fileType = "Continuous" if filepath.split('.')[-1] == "csv" else "Integrated"
        pollutant = filepath.split('\\')[-1].split('_')[0] if fileType == "Continuous" else filepath.split('\\')[-1].split('_')[-1].split('-')[0]
        year = filepath.split('\\')[-1].split('_')[1].split('.')[0] if fileType == "Continuous" else filepath.split('\\')[-1].split('_')[0]

        # for whatever reason, the NAPS data doesn't doesn't download properly
        # tries multiple times to download until expected filesize matches downloaded size
        # TODO: possible infinite while loop
        download_size = -1
        size = int(requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, stream=True).headers["Content-Length"])
        with tqdm(total=size, desc=f"{fileType}-{pollutant}-{year} Downloading") as pbar:
            while abs(download_size - size) > 1: # last byte of data isn't important I think ...
                headers = {"User-Agent": "Mozilla/5.0", "Range": f"bytes={download_size+1}-{size}"}
                with requests.get(url, headers=headers, stream=True, timeout=None) as r:
                    with open(filepath, "ab") as f:
                        shutil.copyfileobj(r.raw, f, 1024)
                        f.flush()
                        os.fsync(f.fileno())
                download_size = os.path.getsize(filepath)
                pbar.update(download_size)

        # if file is a zip file, will unzip contents to the same directory and return directory name
        if filepath.split('.')[-1] == "zip":
            return self.__unzip_file(filepath)
    
    def __unzip_file(self, filepath: str) -> str:

        # if file is a zip file, returns name of unzipped directory

        zip_file = zipfile.ZipFile(filepath, 'r')
        zip_name = [info.filename for info in zip_file.infolist() if info.is_dir()][0][:-1]
        zip_dir_path = f"{os.path.dirname(os.path.realpath(__file__))}\\{zip_name}"
        zip_file.extractall(os.path.dirname(os.path.realpath(__file__)))
            
        return zip_dir_path

    def __delete_files(self, filepath: str, zip_dir_path: str) -> None:

        # deletes file and any existing unzipped directory

        if os.path.exists(filepath):
            os.remove(filepath)
            
        if zip_dir_path is not None and os.path.exists(zip_dir_path):
            shutil.rmtree(zip_dir_path)



if __name__ == "__main__":

    HOST = "192.168.200.72"
    DATABASE = "postgres"
    USER = "socaar_reader"
    PASSWORD = "wallberg123"

    naps = NAPS_Pollutant_Concentrations(HOST, DATABASE, USER, PASSWORD)
    naps.update_data()